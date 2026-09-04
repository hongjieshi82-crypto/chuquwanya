from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


CITY_CODES = {
    "北京": "beijing", "上海": "shanghai", "杭州": "hangzhou", "深圳": "shenzhen",
    "青岛": "qingdao", "南京": "nanjing", "武汉": "wuhan", "成都": "chengdu",
    "西安": "xian", "长沙": "changsha", "重庆": "chongqing", "厦门": "xiamen",
    "天津": "tianjin", "烟台": "yantai", "广州": "guangzhou", "合肥": "hefei",
    "济南": "jinan", "昆明": "kunming",
}

BUDGETS = {
    "当天": (200, 400),
    "周末游": (700, 1300),
    "小长假": (1100, 2000),
}


def budget_label(duration: str, value: int) -> str:
    low, comfortable = BUDGETS[duration]
    if value <= low:
        return "划算出行"
    if value <= comfortable:
        return "舒服躺玩"
    return "品质享受"


def party_range(value: str) -> tuple[int, int]:
    values = value.split("|")
    sizes = []
    for item in values:
        if item == "多人":
            sizes.extend([3, 4])
        else:
            match = re.search(r"\d+", item)
            if match:
                sizes.append(int(match.group()))
    return (min(sizes or [1]), max(sizes or [4]))


def split_schedule(value: str) -> list[str]:
    parts = [part.strip() for part in re.split(r"[；\n]", value or "") if part.strip()]
    return parts or [value.strip()] if value and value.strip() else []


def sql_string(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_json(value: list[str]) -> str:
    return sql_string(json.dumps(value, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--project", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()

    project = args.project.resolve()
    workbook = load_workbook(args.workbook, data_only=False)
    sheet = workbook["完整行程方案"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}

    image_by_row = {}
    for image in sheet._images:
        row = image.anchor._from.row + 1
        if row <= 1:
            continue
        image_by_row[row] = image

    image_dir = project / "apps/client/public/media/itineraries"
    image_dir.mkdir(parents=True, exist_ok=True)
    plans = []
    statements = []

    for row in range(2, sheet.max_row + 1):
        source_id = sheet.cell(row, headers["编号"]).value
        city = sheet.cell(row, headers["城市"]).value
        if not source_id or city not in CITY_CODES:
            continue

        title = str(sheet.cell(row, headers["玩法名称"]).value or "").strip()
        summary = str(sheet.cell(row, headers["一句话简介"]).value or "").strip()
        party = str(sheet.cell(row, headers["适用人数"]).value or "").strip()
        duration = str(sheet.cell(row, headers["出游时长"]).value or "").strip()
        days = int(sheet.cell(row, headers["具体天数"]).value or 1)
        budget = int(sheet.cell(row, headers["人均参考总价（元）"]).value or 0)
        mood = str(sheet.cell(row, headers["心情"]).value or "").strip()
        surprise = str(sheet.cell(row, headers["惊喜程度"]).value or "").strip()
        tags = [value.strip() for value in str(sheet.cell(row, headers["玩法标签（可多选）"]).value or "").split("|") if value.strip()]
        pois = [value.strip() for value in str(sheet.cell(row, headers["主要POI"]).value or "").split("|") if value.strip()]
        included = [value.strip() for value in str(sheet.cell(row, headers["费用包含"]).value or "").split("|") if value.strip()]
        schedule = str(sheet.cell(row, headers["行程安排"]).value or "").strip()
        notice = str(sheet.cell(row, headers["注意事项"]).value or "").strip()
        min_party, max_party = party_range(party)

        cover = None
        image = image_by_row.get(row)
        if image is not None:
            suffix = ".png" if image.format == "png" else ".jpg"
            filename = f"{source_id}{suffix}"
            (image_dir / filename).write_bytes(image._data())
            cover = f"/media/itineraries/{filename}"

        record = {
            "sourceId": source_id,
            "city": city,
            "cityCode": CITY_CODES[city],
            "title": title,
            "summary": summary,
            "partyOptions": party.split("|"),
            "travelDuration": duration,
            "daysCount": days,
            "perPersonBudgetYuan": budget,
            "budgetLabel": budget_label(duration, budget),
            "mood": mood,
            "surpriseLevel": surprise,
            "playTags": tags,
            "poiNames": pois,
            "includedCosts": included,
            "itineraryText": schedule,
            "tips": [notice] if notice else [],
            "status": "review",
            "coverImageUri": cover,
        }
        plans.append(record)

        activity_id = 600000 + len(plans)
        mood_tags = tags + [record["budgetLabel"], duration, f"{surprise}度惊喜"]
        quality_issues = ["批量初稿待核验", "缺少精确坐标与开放时间"]
        steps = split_schedule(schedule)
        primary_poi = pois[0] if pois else city
        category = tags[0] if tags else "城市漫游"
        statements.append(
            "(" + ",".join([
                str(activity_id), f"(SELECT id FROM cities WHERE name={sql_string(city)} LIMIT 1)",
                sql_string(title), sql_string(summary), sql_string(schedule), sql_string(category), sql_string(mood),
                sql_json(mood_tags), sql_string("either"), sql_string("unknown"), sql_string("unknown"), sql_string("unknown"),
                "NULL", "NULL", "NULL", sql_string("unknown"), "NULL", sql_string("review"), "55", sql_json(quality_issues),
                sql_string("itinerary_workbook"), "NULL", sql_string(source_id), sql_json([duration]), "45",
                str(min_party), str(max_party), str(days * 480), str(budget), "0", sql_string("待补充"), sql_string(primary_poi),
                "NULL", "NULL", "NULL", sql_string(cover), sql_json(steps), sql_json([notice] if notice else []), sql_string("#C9FF62"), "TRUE",
            ]) + ")"
        )

    assets_dir = project / "apps/api/assets"
    assets_dir.mkdir(parents=True, exist_ok=True)
    (assets_dir / "itinerary-plans.json").write_text(json.dumps(plans, ensure_ascii=False, indent=2), encoding="utf-8")

    sql = """-- Generated from 18-city itinerary workbook. Imported as review-only content.
DELETE FROM activities WHERE source_type = 'itinerary_workbook';
INSERT INTO activities (
  id,city_id,title,summary,description,category,mood,mood_tags,environment,
  rain_friendly,heat_sensitive,wind_sensitive,weather_notes,last_verified_at,opening_hours,
  reservation_required,reservation_url,content_status,content_score,quality_issues,source_type,
  source_url,place_key,suitable_periods,source_confidence,min_party_size,max_party_size,
  duration_minutes,budget_yuan,city_distance_km,district,address,latitude,longitude,
  navigation_url,cover_image,steps,tips,accent_color,is_active
) VALUES
""" + ",\n".join(statements) + ";\n"
    migration = project / "database/migrations/033_import_18_city_itineraries.sql"
    migration.write_text(sql, encoding="utf-8")
    print(json.dumps({"plans": len(plans), "images": len(image_by_row), "json": str(assets_dir / "itinerary-plans.json"), "migration": str(migration)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
